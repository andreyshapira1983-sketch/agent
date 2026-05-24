"""Tests for tools/builtins/xlsx_writer.py — XlsxWriterTool."""

from __future__ import annotations

from pathlib import Path

import pytest

from tools.builtins.xlsx_writer import XlsxWriterTool


def _load_xlsx(path: Path):
    from openpyxl import load_workbook
    return load_workbook(str(path), data_only=True)


# ════════════════════════════════════════════════════════════════════
# Validation
# ════════════════════════════════════════════════════════════════════

class TestValidation:

    def test_missing_file_path(self):
        result = XlsxWriterTool().execute(sheets=[{"name": "x"}])
        assert not result.success

    def test_wrong_extension(self, tmp_path):
        result = XlsxWriterTool().execute(
            file_path=str(tmp_path / "data.csv"),
            sheets=[{"name": "x"}],
        )
        assert not result.success
        assert ".xlsx" in result.error

    def test_empty_sheets(self, tmp_path):
        result = XlsxWriterTool().execute(
            file_path=str(tmp_path / "data.xlsx"),
            sheets=[],
        )
        assert not result.success

    def test_refuses_existing_file_without_overwrite(self, tmp_path):
        out = tmp_path / "data.xlsx"
        out.write_bytes(b"existing")
        result = XlsxWriterTool().execute(
            file_path=str(out),
            sheets=[{"name": "x", "rows": [[1]]}],
        )
        assert not result.success


# ════════════════════════════════════════════════════════════════════
# Happy path
# ════════════════════════════════════════════════════════════════════

class TestCreate:

    def test_simple_sheet(self, tmp_path):
        out = tmp_path / "data.xlsx"
        result = XlsxWriterTool().execute(
            file_path=str(out),
            sheets=[{
                "name": "Q3",
                "headers": ["Region", "Revenue"],
                "rows": [["EU", 100], ["NA", 200]],
            }],
        )
        assert result.success
        assert out.exists()
        assert result.output["rows"] == 2

        wb = _load_xlsx(out)
        assert "Q3" in wb.sheetnames
        ws = wb["Q3"]
        assert [c.value for c in ws[1]] == ["Region", "Revenue"]
        assert [c.value for c in ws[2]] == ["EU", 100]
        assert [c.value for c in ws[3]] == ["NA", 200]

    def test_multiple_sheets(self, tmp_path):
        out = tmp_path / "data.xlsx"
        result = XlsxWriterTool().execute(
            file_path=str(out),
            sheets=[
                {"name": "First",  "headers": ["a"], "rows": [[1]]},
                {"name": "Second", "headers": ["b"], "rows": [[2]]},
            ],
        )
        assert result.success
        wb = _load_xlsx(out)
        assert wb.sheetnames == ["First", "Second"]

    def test_invalid_sheet_name_cleaned(self, tmp_path):
        out = tmp_path / "data.xlsx"
        result = XlsxWriterTool().execute(
            file_path=str(out),
            sheets=[{"name": "Bad/Name?With:Chars", "rows": [[1]]}],
        )
        assert result.success
        wb = _load_xlsx(out)
        assert all(
            c not in wb.sheetnames[0] for c in "/?:\\*[]"
        ), f"unexpected chars in {wb.sheetnames[0]!r}"

    def test_overwrite(self, tmp_path):
        out = tmp_path / "data.xlsx"
        out.write_bytes(b"old")
        result = XlsxWriterTool().execute(
            file_path=str(out),
            sheets=[{"name": "S", "rows": [[1, 2]]}],
            overwrite=True,
        )
        assert result.success


# ════════════════════════════════════════════════════════════════════
# Limits
# ════════════════════════════════════════════════════════════════════

class TestLimits:

    def test_too_many_rows(self, tmp_path):
        out = tmp_path / "data.xlsx"
        rows = [[i] for i in range(60_000)]
        result = XlsxWriterTool().execute(
            file_path=str(out),
            sheets=[{"name": "S", "rows": rows}],
        )
        assert not result.success
        assert "rows" in result.error.lower()

    def test_long_cell_clipped(self, tmp_path):
        out = tmp_path / "data.xlsx"
        long_cell = "x" * 40_000
        result = XlsxWriterTool().execute(
            file_path=str(out),
            sheets=[{"name": "S", "rows": [[long_cell]]}],
        )
        assert result.success
        wb = _load_xlsx(out)
        v = wb["S"].cell(row=1, column=1).value
        assert isinstance(v, str)
        assert len(v) < 40_000
