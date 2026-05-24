"""
tools/builtins/xlsx_writer.py — Write Excel (.xlsx) spreadsheets.

Companion to XlsxReaderTool. Lets the agent emit tabular data.

API:
    sheets = [
        {
            "name": "Q3",
            "headers": ["Region", "Revenue", "Growth"],
            "rows": [
                ["EU", 12500, 0.12],
                ["NA", 18430, 0.08],
            ],
        },
        ...
    ]

Safety:
    is_destructive=True; refuses to overwrite unless `overwrite=true`.

Caveats:
    The writer aims for "good enough" deliverables. It does NOT support
    formulas, formatting, charts, or merged cells — those belong to a
    later, more specialised tool.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

from ..base import ToolBase, ToolResult, ToolSpec


_MAX_SHEETS = 20
_MAX_ROWS = 50_000
_MAX_COLS = 200
_MAX_CELL_CHARS = 32_000   # Excel limit is ~32,767


class XlsxWriterTool(ToolBase):
    """Create a fresh `.xlsx` file from a list of sheet dicts."""

    @property
    def spec(self) -> ToolSpec:
        return ToolSpec(
            name="xlsx_writer",
            description=(
                "Создаёт Excel-файл (.xlsx) из списка листов. "
                "Каждый лист — dict {name, headers, rows}. "
                "Пишет на диск — destructive."
            ),
            parameters={
                "file_path": "str — путь к итоговому .xlsx",
                "sheets":    "list[dict] — листы {name, headers, rows}",
                "overwrite": "bool (optional, default false)",
            },
            requires_approval=False,
            is_destructive=True,
        )

    # ────────────────────────────────────────────────────────────────

    def execute(self, **params: Any) -> ToolResult:
        file_path = str(params.get("file_path", "")).strip()
        sheets = params.get("sheets") or []
        overwrite = bool(params.get("overwrite", False))

        if not file_path:
            return self._fail("'file_path' is required")
        if not file_path.lower().endswith(".xlsx"):
            return self._fail("file_path must end with .xlsx")
        if not isinstance(sheets, list) or not sheets:
            return self._fail("'sheets' must be a non-empty list of dicts")
        if len(sheets) > _MAX_SHEETS:
            return self._fail(f"too many sheets: {len(sheets)} (max {_MAX_SHEETS})")

        path = Path(file_path).expanduser()
        if path.exists() and not overwrite:
            return self._fail(f"file already exists: {path} (pass overwrite=true)")

        try:
            path.parent.mkdir(parents=True, exist_ok=True)
        except OSError as exc:
            return self._fail(f"cannot create parent dir: {exc}")

        wb = Workbook()
        # openpyxl always creates one default sheet — remove it.
        default = wb.active
        wb.remove(default)

        used_names: set[str] = set()
        total_rows = 0
        for idx, sheet in enumerate(sheets):
            if not isinstance(sheet, dict):
                return self._fail(f"sheet {idx}: must be dict")
            err, rows_count = self._add_sheet(wb, sheet, idx, used_names)
            if err is not None:
                return self._fail(err)
            total_rows += rows_count

        try:
            wb.save(str(path))
        except OSError as exc:
            return self._fail(f"failed to save: {exc}")

        size = path.stat().st_size
        return self._ok({
            "path":       str(path),
            "sheets":     len(sheets),
            "rows":       total_rows,
            "size_bytes": size,
        })

    # ────────────────────────────────────────────────────────────────

    def _add_sheet(self, wb: Workbook, sheet: dict, idx: int, used_names: set[str]) -> tuple[str | None, int]:
        raw_name = str(sheet.get("name", "")).strip() or f"Sheet{idx + 1}"
        # Excel sheet name rules: 31 chars max, no  \ / ? * [ ] :
        clean_name = _clean_sheet_name(raw_name)
        if clean_name in used_names:
            clean_name = f"{clean_name[:28]}_{idx}"
        used_names.add(clean_name)

        headers = sheet.get("headers") or []
        rows = sheet.get("rows") or []

        if headers and not isinstance(headers, list):
            return f"sheet {idx}: 'headers' must be a list", 0
        if not isinstance(rows, list):
            return f"sheet {idx}: 'rows' must be a list of lists", 0
        if len(rows) > _MAX_ROWS:
            return f"sheet {idx}: too many rows ({len(rows)} > {_MAX_ROWS})", 0

        ws = wb.create_sheet(title=clean_name)

        if headers:
            if len(headers) > _MAX_COLS:
                return f"sheet {idx}: too many columns ({len(headers)} > {_MAX_COLS})", 0
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_idx, value=_clip_cell(header))
                cell.font = Font(bold=True)
            start_row = 2
        else:
            start_row = 1

        max_col_used = len(headers)
        for r_idx, row in enumerate(rows, start=start_row):
            if not isinstance(row, (list, tuple)):
                return f"sheet {idx}: row {r_idx - start_row + 1} must be a list", 0
            if len(row) > _MAX_COLS:
                return f"sheet {idx} row {r_idx}: too many columns", 0
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=_clip_cell(value))
            max_col_used = max(max_col_used, len(row))

        # Reasonable column widths
        for c in range(1, max_col_used + 1):
            ws.column_dimensions[get_column_letter(c)].width = 18

        return None, len(rows)


# ════════════════════════════════════════════════════════════════════
# Helpers
# ════════════════════════════════════════════════════════════════════

_INVALID_SHEET_CHARS = set(r'\/?*[]:')


def _clean_sheet_name(name: str) -> str:
    cleaned = "".join("_" if c in _INVALID_SHEET_CHARS else c for c in name)
    return cleaned[:31] or "Sheet"


def _clip_cell(value: Any) -> Any:
    if isinstance(value, str) and len(value) > _MAX_CELL_CHARS:
        return value[: _MAX_CELL_CHARS - 1] + "…"
    return value
